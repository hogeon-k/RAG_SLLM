from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

try:
    from scripts.generate_goal2_test_workbooks import DEFAULT_OUTPUT_DIR, calculate_sha256
except ModuleNotFoundError:
    from generate_goal2_test_workbooks import DEFAULT_OUTPUT_DIR, calculate_sha256


def verify_goal2_test_workbooks(output_dir: Path = DEFAULT_OUTPUT_DIR) -> list[str]:
    failures: list[str] = []
    main = output_dir / "goal2_regulations_fixture.xlsx"
    duplicate = output_dir / "goal2_regulations_fixture_duplicate.xlsx"
    modified = output_dir / "goal2_regulations_fixture_modified.xlsx"
    corrupted = output_dir / "goal2_corrupted.xlsx"
    manifest = output_dir / "goal2_fixture_manifest.json"

    required = [main, duplicate, modified, corrupted, manifest]
    for path in required:
        _check(path.exists(), f"exists: {path.name}", failures)

    if failures:
        return failures

    workbook = load_workbook(main, data_only=False)
    try:
        expected_order = ["문서정보", "휴가규정", "출장규정", "데이터유형", "장문규정", "빈시트", "내부메모", "시스템자료"]
        _check(workbook.sheetnames == expected_order, "main sheet order matches", failures)
        _check(workbook["내부메모"].sheet_state == "hidden", "hidden sheet state matches", failures)
        _check(workbook["시스템자료"].sheet_state == "veryHidden", "veryHidden sheet state matches", failures)
        _check("A1:F1" in [str(r) for r in workbook["문서정보"].merged_cells.ranges], "document title merged range exists", failures)
        _check("B14:F14" in [str(r) for r in workbook["휴가규정"].merged_cells.ranges], "leave article merged range exists", failures)
        _check(workbook["문서정보"]["B3"].value == "가상 주식회사 업무규정", "document name matches", failures)
        _check(workbook["휴가규정"]["A13"].value == "제8조", "article 제8조 marker exists", failures)
        _check(workbook["휴가규정"]["B13"].value == "연차휴가 신청", "article 제8조 title exists", failures)
        _check("3일 전" in workbook["휴가규정"]["B14"].value, "main leave deadline is 3 days", failures)
        _check(workbook["휴가규정"]["A18"].value == "제8조의2(긴급휴가)", "article 제8조의2 exists", failures)
        _check("80%" in workbook["휴가규정"]["B23"].value, "plain numeric-looking text exists", failures)
        _check(workbook["데이터유형"]["B3"].value == "테스트 문자열", "string value exists", failures)
        _check(workbook["데이터유형"]["B4"].value == 42, "integer value exists", failures)
        _check(workbook["데이터유형"]["B22"].value == "=SUM(B20:B21)", "formula cell exists", failures)
        long_text = "\n".join(str(workbook["장문규정"][f"A{row}"].value or "") for row in range(6, 18))
        _check(len(long_text) >= 1800, "long article has at least 1800 characters", failures)
    finally:
        workbook.close()

    modified_workbook = load_workbook(modified, data_only=False)
    try:
        _check(modified_workbook["문서정보"]["B4"].value == "2.1", "modified version is 2.1", failures)
        _check("5일 전" in modified_workbook["휴가규정"]["B14"].value, "modified leave deadline is 5 days", failures)
    finally:
        modified_workbook.close()

    _check(calculate_sha256(main) == calculate_sha256(duplicate), "main and duplicate SHA-256 match", failures)
    _check(calculate_sha256(main) != calculate_sha256(modified), "main and modified SHA-256 differ", failures)

    try:
        corrupted_workbook = load_workbook(corrupted)
    except Exception:
        corrupted_workbook = None
    finally:
        if "corrupted_workbook" in locals() and corrupted_workbook is not None:
            corrupted_workbook.close()
    _check(corrupted_workbook is None, "corrupted workbook cannot be opened by openpyxl", failures)

    return failures


def _check(condition: bool, label: str, failures: list[str]) -> None:
    status = "PASS" if condition else "FAIL"
    print(f"[{status}] {label}")
    if not condition:
        failures.append(label)


def main() -> int:
    failures = verify_goal2_test_workbooks(DEFAULT_OUTPUT_DIR)
    if failures:
        print("\nVerification failed:")
        for failure in failures:
            print(f"- {failure}")
        return 1
    print("\nPASS: all Goal 2.5 workbook checks succeeded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
