from __future__ import annotations

import hashlib
import json
import shutil
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "test_workbooks"
GENERATOR_VERSION = "1.0.0"
FIXED_CREATED_AT = datetime(2026, 7, 1, 9, 0, 0)


@dataclass(frozen=True)
class GenerationResult:
    output_dir: Path
    main_path: Path
    duplicate_path: Path
    modified_path: Path
    corrupted_path: Path
    manifest_path: Path
    hashes: dict[str, str]


def generate_test_workbooks(output_dir: Path = DEFAULT_OUTPUT_DIR) -> GenerationResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    main_path = output_dir / "goal2_regulations_fixture.xlsx"
    duplicate_path = output_dir / "goal2_regulations_fixture_duplicate.xlsx"
    modified_path = output_dir / "goal2_regulations_fixture_modified.xlsx"
    corrupted_path = output_dir / "goal2_corrupted.xlsx"
    manifest_path = output_dir / "goal2_fixture_manifest.json"

    _build_main_workbook(main_path)
    shutil.copy2(main_path, duplicate_path)
    shutil.copy2(main_path, modified_path)
    _modify_workbook(modified_path)
    corrupted_path.write_bytes(b"This is not a valid XLSX file. It is a Goal 2.5 corrupted fixture.\n")

    hashes = {
        "main": calculate_sha256(main_path),
        "duplicate": calculate_sha256(duplicate_path),
        "modified": calculate_sha256(modified_path),
        "corrupted": calculate_sha256(corrupted_path),
    }
    if hashes["main"] != hashes["duplicate"]:
        raise RuntimeError("main and duplicate workbook hashes must match.")
    if hashes["main"] == hashes["modified"]:
        raise RuntimeError("main and modified workbook hashes must differ.")

    manifest = _manifest(output_dir, hashes)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return GenerationResult(output_dir, main_path, duplicate_path, modified_path, corrupted_path, manifest_path, hashes)


def calculate_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _build_main_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "가상 주식회사 업무규정"
    workbook.properties.subject = "RAG_SLLM Goal 2 테스트"
    workbook.properties.creator = "RAG_SLLM Test Generator"
    workbook.properties.description = "실제 회사 정보가 아닌 기능 검증용 가상 데이터"
    workbook.properties.created = FIXED_CREATED_AT
    workbook.properties.modified = FIXED_CREATED_AT

    _sheet_document_info(workbook.create_sheet("문서정보"))
    _sheet_leave(workbook.create_sheet("휴가규정"))
    _sheet_travel(workbook.create_sheet("출장규정"))
    _sheet_data_types(workbook.create_sheet("데이터유형"))
    _sheet_long_policy(workbook.create_sheet("장문규정"))
    _sheet_empty(workbook.create_sheet("빈시트"))
    memo = workbook.create_sheet("내부메모")
    _sheet_hidden_memo(memo)
    memo.sheet_state = "hidden"
    system = workbook.create_sheet("시스템자료")
    _sheet_system(system)
    system.sheet_state = "veryHidden"

    workbook.save(path)
    workbook.close()


def _modify_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    workbook.properties.modified = datetime(2026, 7, 2, 9, 0, 0)
    workbook["문서정보"]["B4"] = "2.1"
    workbook["휴가규정"]["B14"] = "연차휴가는 사용 예정일 5일 전까지 신청해야 한다."
    workbook.save(path)
    workbook.close()


def _sheet_document_info(ws) -> None:
    _title(ws, "A1:F1", "가상 주식회사 업무규정")
    rows = [
        ("A3", "문서명", "B3", "가상 주식회사 업무규정"),
        ("A4", "버전", "B4", "2.0"),
        ("A5", "시행일", "B5", datetime(2026, 7, 1)),
        ("A6", "개정일", "B6", datetime(2026, 6, 15)),
        ("A7", "담당 부서", "B7", "가상 경영지원팀"),
        ("A8", "문서 구분", "B8", "테스트용 가상 규정"),
    ]
    for label_cell, label, value_cell, value in rows:
        ws[label_cell] = label
        ws[value_cell] = value
    ws["B5"].number_format = "yyyy-mm-dd"
    ws["B6"].number_format = "yyyy-mm-dd"
    ws["A10"] = "이 문서는 실제 회사 규정이 아닌 RAG_SLLM 기능 검증용 더미 문서입니다."
    ws["A12"] = "빈 행과 빈 셀 처리 확인을 위해 일부 영역은 의도적으로 비워 둡니다."
    _format_basic(ws, 6)


def _sheet_leave(ws) -> None:
    _title(ws, "A1:F1", "휴가규정")
    _section(ws, "A3:F3", "제1장 총칙")
    _section(ws, "A5:F5", "제1절 휴가 신청")
    _row(ws, 7, "제1조", "목적")
    _row(ws, 8, "①", "이 규정은 임직원의 휴가 신청 절차와 승인 기준을 정하는 것을 목적으로 한다.")
    _row(ws, 10, "제2조(적용범위)", "")
    _merge_text(ws, "B11:F11", "이 규정은 가상 주식회사에 근무하는 모든 임직원에게 적용한다.")
    _row(ws, 13, "제8조", "연차휴가 신청")
    _row(ws, 14, "①", "연차휴가는 사용 예정일 3일 전까지 신청해야 한다.")
    _row(ws, 15, "②", "신청자는 가상 인사시스템에서 휴가신청서를 작성해야 한다.")
    _row(ws, 16, "③", "부서장의 승인이 완료된 이후 휴가를 사용할 수 있다.")
    _merge_text(ws, "A18:F18", "제8조의2(긴급휴가)")
    _row(ws, 19, "①", "질병이나 가족 사고 등 긴급한 사유가 있으면 당일 신청할 수 있다.")
    _row(ws, 20, "②", "당일 신청자는 업무 시작 전까지 담당자에게 사유를 알려야 한다.")
    ws["A23"] = "참고"
    ws["B23"] = "신청 완료율은 80%이며 처리 시간은 평균 3일이다."
    ws["A25"] = "숨김 행 참고"
    ws["B25"] = "이 행은 숨김 행 처리 검증용이다."
    ws.row_dimensions[25].hidden = True
    ws["F24"] = "숨김 열 처리 검증용"
    ws.column_dimensions["F"].hidden = True
    _format_basic(ws, 6)


def _sheet_travel(ws) -> None:
    _title(ws, "A1:G1", "출장규정")
    _section(ws, "A3:G3", "제1장 국내출장")
    _section(ws, "A5:G5", "제1절 출장비")
    ws["A7"] = "제2조(출장 신청)"
    ws["A8"] = "출장자는 출발일 2일 전까지 출장신청서를 제출해야 한다."
    ws["A10"] = "제3조"
    _merge_text(ws, "B10:G10", "제출 서류")
    headers = ["구분", "필수 서류", "제출 기한", "담당 부서", "최대 금액", "승인 필요", "비고"]
    for column, header in enumerate(headers, 1):
        ws.cell(row=12, column=column, value=header)
    rows = [
        ["국내출장", "출장신청서", "출발 2일 전", "가상 총무팀", 150000, True, "사전 승인"],
        ["숙박비", "영수증", "복귀 후 3일 이내", "가상 회계팀", 100000, True, "실비 정산"],
        ["교통비", "영수증", "복귀 후 3일 이내", "가상 회계팀", 80000, False, "대중교통 기준"],
    ]
    for row_index, row_values in enumerate(rows, 13):
        for column, value in enumerate(row_values, 1):
            ws.cell(row=row_index, column=column, value=value)
    ws["A17"] = "제4조(정산)"
    ws["A18"] = "출장 종료 후 3일 이내에 영수증과 출장보고서를 제출해야 한다."
    _format_basic(ws, 7)


def _sheet_data_types(ws) -> None:
    _title(ws, "A1:D1", "데이터유형")
    values = [
        ("문자열", "테스트 문자열"),
        ("정수", 42),
        ("실수", 1234.567),
        ("음수", -25),
        ("백분율", 0.8),
        ("통화", 150000),
        ("Boolean True", True),
        ("Boolean False", False),
        ("date", datetime(2026, 7, 1)),
        ("datetime", datetime(2026, 7, 1, 9, 30, 0)),
        ("time", time(9, 30, 0)),
        ("다중 줄 문자열", "첫 번째 줄\n두 번째 줄"),
        ("공백 정리", "  앞뒤 및   연속   공백이   있는 문자열  "),
        ("한글영문혼합", "가상Policy ABC-123"),
        ("오류 값", "#DIV/0!"),
        ("빈 셀", None),
        ("공백 문자열", "   "),
    ]
    for row_index, (label, value) in enumerate(values, 3):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
    ws["B7"].number_format = "0%"
    ws["B8"].number_format = '#,##0"원"'
    ws["B11"].number_format = "yyyy-mm-dd"
    ws["B12"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws["B13"].number_format = "hh:mm:ss"
    ws["A22"] = "합계 수식"
    ws["B20"] = 100
    ws["B21"] = 200
    ws["B22"] = "=SUM(B20:B21)"
    _format_basic(ws, 4)


def _sheet_long_policy(ws) -> None:
    _title(ws, "A1:F1", "장문규정")
    _section(ws, "A3:F3", "제1장 정보보호")
    ws["A5"] = "제10조(정보보호 기본원칙)"
    paragraphs = [
        "① 개인정보는 업무 목적에 필요한 범위에서만 수집하며 가상의 승인 절차를 거친다. 수집 목적, 보관 기간, 접근 가능 역할을 문서화하고 임의 저장소에 복사하지 않는다.",
        "② 비밀번호는 정기적으로 변경하고 다른 사람에게 공유하지 않는다. 테스트 계정도 개인 계정처럼 관리하며 화면 캡처나 메신저로 전달하지 않는다.",
        "③ 접근권한은 담당 업무와 역할에 맞게 최소 범위로 부여한다. 임시 권한은 시작일과 종료일을 기록하고 목적이 끝나면 즉시 회수한다.",
        "④ 로그 기록은 이상 징후 확인을 위해 보관하며 임의로 삭제하지 않는다. 점검자는 필요한 항목만 열람하고 원문 로그 전체를 불필요하게 반출하지 않는다.",
        "⑤ 외부 반출이 필요한 자료는 가상 보안 담당자의 확인을 받은 뒤 처리한다. 반출 사유, 대상 파일, 수신자, 회수 계획을 간단히 기록한다.",
        "⑥ 백업 파일은 지정된 저장소에 보관하고 무단 복사본을 만들지 않는다. 백업 복원 테스트는 별도 검증 환경에서 수행하고 결과만 요약한다.",
        "⑦ 사고가 의심되는 경우 즉시 보안 신고 채널로 보고한다. 보고자는 추정 원인을 단정하지 않고 발견 시각, 영향 범위, 조치 현황을 구분해 작성한다.",
        "⑧ 교육 자료와 안내 문서는 최신 버전을 확인한 뒤 사용한다. 오래된 지침이 발견되면 담당 부서에 개정 여부를 문의하고 자체 수정하지 않는다.",
        "⑨ 테스트 환경의 계정과 운영 환경의 계정은 분리하여 관리한다. 테스트 데이터에는 실제 고객명, 주민등록번호, 계좌번호 같은 민감 정보를 넣지 않는다.",
        "⑩ 본 조의 예시는 기능 검증을 위한 가상 문구이며 실제 규정이 아니다. 문구는 청크 분할과 셀 위치 검증을 위해 충분히 길게 작성되었다.",
        "⑪ 추가 점검 항목은 월별 보안 회의에서 검토하고 필요한 경우 개정한다. 검토 결과는 버전과 시행일을 포함해 별도 변경 이력으로 남긴다.",
        "⑫ 임시 권한은 사용 목적이 종료되면 지체 없이 회수한다. 회수 확인자는 처리 시각과 확인자를 기록하고 미회수 항목은 다음 영업일 전까지 재점검한다.",
    ]
    paragraphs = [
        paragraph
        + " 이 문장은 장문 청크 분할과 셀 범위 보존을 검증하기 위한 가상의 보충 설명이다."
        + " 각 문단은 독립된 셀 위치를 유지하며 실제 회사 정보나 개인정보를 포함하지 않는다."
        for paragraph in paragraphs
    ]
    for index, text in enumerate(paragraphs, 6):
        _merge_text(ws, f"A{index}:F{index}", text)
    ws["A20"] = "제11조(보안사고 보고)"
    ws["A21"] = "보안사고를 인지한 임직원은 지체 없이 가상 보안 담당자에게 보고해야 한다."
    _format_basic(ws, 6)


def _sheet_empty(ws) -> None:
    ws["A1"] = None


def _sheet_hidden_memo(ws) -> None:
    ws["A1"] = "내부 테스트 메모"
    ws["A2"] = "이 시트는 기본 설정에서 청크 생성 대상에서 제외되어야 한다."


def _sheet_system(ws) -> None:
    ws["A1"] = "시스템 테스트 데이터"
    ws["A2"] = "이 시트는 기본 설정에서 청크 생성 대상에서 제외되어야 한다."


def _title(ws, range_ref: str, text: str) -> None:
    ws.merge_cells(range_ref)
    cell = ws[range_ref.split(":")[0]]
    cell.value = text
    cell.font = Font(bold=True, size=15)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="D9EAD3")


def _section(ws, range_ref: str, text: str) -> None:
    ws.merge_cells(range_ref)
    cell = ws[range_ref.split(":")[0]]
    cell.value = text
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDEBF7")


def _row(ws, row: int, marker: str, text: str) -> None:
    ws.cell(row=row, column=1, value=marker)
    if text:
        _merge_text(ws, f"B{row}:F{row}", text)


def _merge_text(ws, range_ref: str, text: str) -> None:
    ws.merge_cells(range_ref)
    cell = ws[range_ref.split(":")[0]]
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _format_basic(ws, max_col: int) -> None:
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24


def _manifest(output_dir: Path, hashes: dict[str, str]) -> dict:
    sheet_order = ["문서정보", "휴가규정", "출장규정", "데이터유형", "장문규정", "빈시트", "내부메모", "시스템자료"]
    return {
        "generator_version": GENERATOR_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "output_dir": str(output_dir),
        "files": {
            "main": {"name": "goal2_regulations_fixture.xlsx", "sha256": hashes["main"]},
            "duplicate": {"name": "goal2_regulations_fixture_duplicate.xlsx", "sha256": hashes["duplicate"]},
            "modified": {"name": "goal2_regulations_fixture_modified.xlsx", "sha256": hashes["modified"]},
            "corrupted": {"name": "goal2_corrupted.xlsx", "sha256": hashes["corrupted"]},
        },
        "hash_checks": {
            "main_duplicate_same": hashes["main"] == hashes["duplicate"],
            "main_modified_different": hashes["main"] != hashes["modified"],
        },
        "sheet_order": sheet_order,
        "sheet_states": {
            "문서정보": "visible",
            "휴가규정": "visible",
            "출장규정": "visible",
            "데이터유형": "visible",
            "장문규정": "visible",
            "빈시트": "visible",
            "내부메모": "hidden",
            "시스템자료": "veryHidden",
        },
        "merged_ranges": ["문서정보!A1:F1", "휴가규정!A1:F1", "휴가규정!B14:F14", "출장규정!A1:G1", "장문규정!A6:F17"],
        "expected_articles": ["제1조", "제2조", "제8조", "제8조의2", "제10조", "제11조"],
        "hidden_sheets": ["내부메모", "시스템자료"],
        "formula_cells": {"데이터유형!B22": "=SUM(B20:B21)"},
        "long_article": {"sheet": "장문규정", "article": "제10조", "minimum_chars": 1800},
        "modified_changes": {
            "문서정보!B4": "2.1",
            "휴가규정!B14": "연차휴가는 사용 예정일 5일 전까지 신청해야 한다.",
        },
        "corrupted_note": "goal2_corrupted.xlsx has an .xlsx extension but is not a valid XLSX archive.",
        "formula_cache_note": "openpyxl does not calculate formulas; generated formula cells may not have cached values.",
    }


def main() -> int:
    result = generate_test_workbooks(DEFAULT_OUTPUT_DIR)
    print(f"Generated workbooks in: {result.output_dir}")
    for name, digest in result.hashes.items():
        print(f"{name}: {digest}")
    print(f"Manifest: {result.manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
