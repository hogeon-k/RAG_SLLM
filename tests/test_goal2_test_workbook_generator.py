from __future__ import annotations

from openpyxl import load_workbook

from scripts.generate_goal2_test_workbooks import calculate_sha256, generate_test_workbooks
from scripts.verify_goal2_test_workbooks import verify_goal2_test_workbooks


def test_generator_creates_expected_workbooks(tmp_path) -> None:
    result = generate_test_workbooks(tmp_path)

    assert result.main_path.exists()
    assert result.duplicate_path.exists()
    assert result.modified_path.exists()
    assert result.corrupted_path.exists()
    assert result.manifest_path.exists()


def test_generator_hash_relationships(tmp_path) -> None:
    result = generate_test_workbooks(tmp_path)

    assert calculate_sha256(result.main_path) == calculate_sha256(result.duplicate_path)
    assert calculate_sha256(result.main_path) != calculate_sha256(result.modified_path)


def test_generator_sheet_state_and_merged_ranges(tmp_path) -> None:
    result = generate_test_workbooks(tmp_path)
    workbook = load_workbook(result.main_path)
    try:
        assert workbook.sheetnames == ["문서정보", "휴가규정", "출장규정", "데이터유형", "장문규정", "빈시트", "내부메모", "시스템자료"]
        assert workbook["내부메모"].sheet_state == "hidden"
        assert workbook["시스템자료"].sheet_state == "veryHidden"
        assert "A1:F1" in [str(item) for item in workbook["휴가규정"].merged_cells.ranges]
        assert workbook["휴가규정"]["A13"].value == "제8조"
        assert workbook["휴가규정"]["B13"].value == "연차휴가 신청"
        assert "3일 전" in workbook["휴가규정"]["B14"].value
        assert workbook["휴가규정"]["A18"].value == "제8조의2(긴급휴가)"
    finally:
        workbook.close()


def test_generator_corrupted_file_fails_open(tmp_path) -> None:
    result = generate_test_workbooks(tmp_path)

    try:
        workbook = load_workbook(result.corrupted_path)
    except Exception:
        workbook = None
    finally:
        if "workbook" in locals() and workbook is not None:
            workbook.close()

    assert workbook is None


def test_verify_script_passes_generated_files(tmp_path) -> None:
    generate_test_workbooks(tmp_path)

    assert verify_goal2_test_workbooks(tmp_path) == []
