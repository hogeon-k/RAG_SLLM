from __future__ import annotations

from datetime import date

import pytest
from openpyxl import Workbook

from app.services.excel_parser_service import ExcelParserService
from app.services.exceptions import DocumentExtractionError


def _workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "규정"
    sheet["A1"] = "문자"
    sheet["B1"] = 10
    sheet["C1"] = True
    sheet["D1"] = date(2026, 1, 1)
    sheet["A2"] = "  공백   정리  "
    workbook.save(path)
    workbook.close()
    return path


def test_extracts_basic_cell_types(tmp_path) -> None:
    path = _workbook(tmp_path / "basic.xlsx")

    parsed = ExcelParserService().parse(path, "DOC-1")

    values = {cell.coordinate: cell for cell in parsed.cells}
    assert values["A1"].text_value == "문자"
    assert values["B1"].value_type == "integer"
    assert values["C1"].text_value == "True"
    assert values["D1"].value_type == "date"


def test_empty_cells_excluded_and_spaces_normalized(tmp_path) -> None:
    path = _workbook(tmp_path / "basic.xlsx")

    parsed = ExcelParserService().parse(path, "DOC-1")

    values = {cell.coordinate: cell for cell in parsed.cells}
    assert "E1" not in values
    assert values["A2"].text_value == "공백 정리"


def test_sheet_order_and_state_preserved(tmp_path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "첫번째"
    first["A1"] = "내용"
    second = workbook.create_sheet("숨김")
    second.sheet_state = "hidden"
    second["A1"] = "숨김 내용"
    path = tmp_path / "sheets.xlsx"
    workbook.save(path)
    workbook.close()

    parsed = ExcelParserService().parse(path, "DOC-1")

    assert [sheet.sheet_name for sheet in parsed.sheets] == ["첫번째", "숨김"]
    assert parsed.sheets[1].sheet_state == "hidden"
    assert parsed.skipped_hidden_sheet_count == 1
    assert all(cell.sheet_name != "숨김" for cell in parsed.cells)


def test_hidden_sheet_included_when_enabled(tmp_path) -> None:
    workbook = Workbook()
    first = workbook.active
    first["A1"] = "보임"
    hidden = workbook.create_sheet("숨김")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "숨김"
    path = tmp_path / "hidden.xlsx"
    workbook.save(path)
    workbook.close()

    parsed = ExcelParserService(include_hidden_sheets=True).parse(path, "DOC-1")

    assert any(cell.sheet_name == "숨김" for cell in parsed.cells)


def test_cell_limit_rejected(tmp_path) -> None:
    path = _workbook(tmp_path / "basic.xlsx")

    with pytest.raises(DocumentExtractionError):
        ExcelParserService(max_extracted_cells=1).parse(path, "DOC-1")


def test_merged_cell_anchor_saved_once_with_range(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("B2:F2")
    sheet["B2"] = "병합 제목"
    path = tmp_path / "merged.xlsx"
    workbook.save(path)
    workbook.close()

    parsed = ExcelParserService().parse(path, "DOC-1")

    assert len(parsed.cells) == 1
    assert parsed.cells[0].coordinate == "B2"
    assert parsed.cells[0].merged_range == "B2:F2"
    assert parsed.cells[0].is_merged_anchor


def test_formula_preserved_without_modifying_file(tmp_path) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["A3"] = "=SUM(A1:A2)"
    workbook.save(path)
    before = path.read_bytes()
    workbook.close()

    parsed = ExcelParserService().parse(path, "DOC-1")

    formula_cell = next(cell for cell in parsed.cells if cell.coordinate == "A3")
    assert formula_cell.formula == "=SUM(A1:A2)"
    assert formula_cell.text_value == "=SUM(A1:A2)"
    assert path.read_bytes() == before

