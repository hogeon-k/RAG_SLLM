from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from app.models.document import ParsedCell, ParsedSheet, ParsedWorkbook
from app.services.exceptions import DocumentExtractionError


class ExcelParserService:
    def __init__(self, max_extracted_cells: int = 200000, include_hidden_sheets: bool = False) -> None:
        self._max_extracted_cells = max_extracted_cells
        self._include_hidden_sheets = include_hidden_sheets

    def parse(self, file_path: Path, document_id: str) -> ParsedWorkbook:
        if not file_path.exists():
            raise DocumentExtractionError("저장된 원본 파일을 찾을 수 없습니다.")

        workbook = load_workbook(file_path, read_only=False, data_only=False, keep_links=False)
        values_workbook = load_workbook(file_path, read_only=False, data_only=True, keep_links=False)
        sheets: list[ParsedSheet] = []
        cells: list[ParsedCell] = []
        skipped_hidden_sheet_count = 0
        created_at = datetime.now().replace(microsecond=0)

        try:
            if not workbook.worksheets:
                raise DocumentExtractionError("엑셀 파일에 시트가 없습니다.")

            for sheet_index, worksheet in enumerate(workbook.worksheets):
                sheet_id = f"{document_id}-S{sheet_index + 1:03d}"
                value_sheet = values_workbook[worksheet.title]
                merged_lookup = _merged_lookup(worksheet)
                visible_for_cells = worksheet.sheet_state == "visible" or self._include_hidden_sheets
                if worksheet.sheet_state != "visible" and not self._include_hidden_sheets:
                    skipped_hidden_sheet_count += 1

                sheet_cells: list[ParsedCell] = []
                if visible_for_cells:
                    for row in worksheet.iter_rows():
                        for cell in row:
                            parsed_cell = _parse_cell(
                                document_id,
                                sheet_id,
                                worksheet.title,
                                cell,
                                value_sheet[cell.coordinate].value,
                                merged_lookup,
                                worksheet.row_dimensions[cell.row].hidden,
                                worksheet.column_dimensions[get_column_letter(cell.column)].hidden,
                            )
                            if parsed_cell:
                                sheet_cells.append(parsed_cell)

                total_count = len(cells) + len(sheet_cells)
                if total_count > self._max_extracted_cells:
                    raise DocumentExtractionError("추출할 셀 수가 설정된 제한을 초과했습니다.")

                sheets.append(
                    ParsedSheet(
                        id=sheet_id,
                        document_id=document_id,
                        sheet_name=worksheet.title,
                        sheet_index=sheet_index,
                        sheet_state=worksheet.sheet_state,
                        max_row=worksheet.max_row,
                        max_column=worksheet.max_column,
                        non_empty_cell_count=len(sheet_cells),
                        merged_range_count=len(worksheet.merged_cells.ranges),
                        created_at=created_at,
                    )
                )
                cells.extend(sheet_cells)
        finally:
            workbook.close()
            values_workbook.close()

        if not cells:
            raise DocumentExtractionError("추출할 수 있는 셀 내용이 없습니다.")

        return ParsedWorkbook(
            document_id=document_id,
            sheets=tuple(sheets),
            cells=tuple(cells),
            skipped_hidden_sheet_count=skipped_hidden_sheet_count,
            warnings=(),
        )


def _parse_cell(
    document_id: str,
    sheet_id: str,
    sheet_name: str,
    cell: Cell,
    cached_value: Any,
    merged_lookup: dict[str, tuple[str, bool]],
    row_hidden: bool,
    column_hidden: bool,
) -> ParsedCell | None:
    merged = merged_lookup.get(cell.coordinate)
    if merged and not merged[1]:
        return None

    raw_value = cell.value
    formula = raw_value if isinstance(raw_value, str) and raw_value.startswith("=") else None
    value_type = "formula" if formula else _value_type(raw_value)
    cached_text = _normalize_value(cached_value)
    text_value = cached_text if formula and cached_text else _normalize_value(raw_value)
    if not text_value:
        return None

    return ParsedCell(
        id=f"{sheet_id}-{cell.coordinate}",
        document_id=document_id,
        sheet_id=sheet_id,
        sheet_name=sheet_name,
        coordinate=cell.coordinate,
        row_index=cell.row,
        column_index=cell.column,
        value_type=value_type,
        text_value=text_value,
        formula=formula,
        cached_value=cached_text if formula else None,
        merged_range=merged[0] if merged else None,
        is_merged_anchor=bool(merged and merged[1]),
        is_hidden=bool(row_hidden or column_hidden),
    )


def _merged_lookup(worksheet) -> dict[str, tuple[str, bool]]:
    lookup: dict[str, tuple[str, bool]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        range_text = str(merged_range)
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor = f"{worksheet.cell(row=min_row, column=min_col).coordinate}"
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                coordinate = worksheet.cell(row=row, column=column).coordinate
                lookup[coordinate] = (range_text, coordinate == anchor)
    return lookup


def _value_type(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return "date"
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    if isinstance(value, str) and value.startswith("#"):
        return "error"
    return "string"


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date | time):
        return value.isoformat()
    if isinstance(value, bool):
        return "True" if value else "False"
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line).strip()
