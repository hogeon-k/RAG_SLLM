from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import datetime

from openpyxl.utils import get_column_letter, range_boundaries

from app.models.document import DocumentChunk, ParsedCell, ParsedWorkbook


SECTION_PATTERN = re.compile(r"^제\s*\d+\s*(장|절)\b")
ARTICLE_PATTERN = re.compile(r"^(제\s*\d+\s*조(?:의\s*\d+)?)(?:\s*\(([^)]+)\))?")
PARAGRAPH_PATTERN = re.compile(r"^([①-⑳]|\d+\.|[가-힣]\.)\s*")


@dataclass(frozen=True)
class RowText:
    sheet_id: str
    sheet_name: str
    row_index: int
    text: str
    cell_refs: tuple[str, ...]
    min_row: int
    min_col: int
    max_row: int
    max_col: int


class ChunkService:
    def __init__(self, max_chars: int = 1500, min_chars: int = 80) -> None:
        self._max_chars = max_chars
        self._min_chars = min_chars

    def create_chunks(self, parsed_workbook: ParsedWorkbook) -> list[DocumentChunk]:
        chunks: list[DocumentChunk] = []
        created_at = datetime.now().replace(microsecond=0)
        for sheet in parsed_workbook.sheets:
            rows = _rows_for_sheet(parsed_workbook.cells, sheet.id)
            sheet_chunks = self._chunks_for_sheet(parsed_workbook.document_id, sheet.id, sheet.sheet_name, rows, created_at)
            chunks.extend(sheet_chunks)
        return chunks

    def _chunks_for_sheet(
        self,
        document_id: str,
        sheet_id: str,
        sheet_name: str,
        rows: list[RowText],
        created_at: datetime,
    ) -> list[DocumentChunk]:
        chunks: list[DocumentChunk] = []
        active_rows: list[RowText] = []
        section: str | None = None
        article: str | None = None
        title: str | None = None

        for row in rows:
            row_section, row_article, row_title, _paragraph = _classify(row.text)
            if row_section:
                section = row_section
            starts_new_article = bool(row_article)
            if starts_new_article and active_rows:
                chunks.extend(self._flush_rows(document_id, sheet_id, sheet_name, active_rows, section, article, title, created_at))
                active_rows = []
            if row_article:
                article = row_article
                title = row_title

            if len(_join_rows(active_rows + [row])) > self._max_chars and active_rows:
                chunks.extend(self._flush_rows(document_id, sheet_id, sheet_name, active_rows, section, article, title, created_at))
                active_rows = []
            active_rows.append(row)

        if active_rows:
            chunks.extend(self._flush_rows(document_id, sheet_id, sheet_name, active_rows, section, article, title, created_at))

        return [
            _with_chunk_index(chunk, index)
            for index, chunk in enumerate(chunks)
            if chunk.content.strip()
        ]

    def _flush_rows(
        self,
        document_id: str,
        sheet_id: str,
        sheet_name: str,
        rows: list[RowText],
        section: str | None,
        article: str | None,
        title: str | None,
        created_at: datetime,
    ) -> list[DocumentChunk]:
        if not rows:
            return []
        split_groups = _split_rows_by_length(rows, self._max_chars)
        chunks: list[DocumentChunk] = []
        for group in split_groups:
            content = _join_rows(group)
            if not content:
                continue
            chunks.append(_make_chunk(document_id, sheet_id, sheet_name, group, section, article, title, created_at, 0))
        return chunks


def _rows_for_sheet(cells: tuple[ParsedCell, ...], sheet_id: str) -> list[RowText]:
    grouped: dict[int, list[ParsedCell]] = {}
    for cell in cells:
        if cell.sheet_id == sheet_id and not cell.is_hidden:
            grouped.setdefault(cell.row_index, []).append(cell)

    rows: list[RowText] = []
    for row_index in sorted(grouped):
        row_cells = sorted(grouped[row_index], key=lambda item: item.column_index)
        parts = [cell.text_value for cell in row_cells if cell.text_value.strip()]
        if not parts:
            continue
        refs = tuple(_cell_ref(cell) for cell in row_cells if cell.text_value.strip())
        min_row, min_col, max_row, max_col = _bounds_for_cells(row_cells)
        rows.append(
            RowText(
                sheet_id=sheet_id,
                sheet_name=row_cells[0].sheet_name,
                row_index=row_index,
                text=" | ".join(parts),
                cell_refs=refs,
                min_row=min_row,
                min_col=min_col,
                max_row=max_row,
                max_col=max_col,
            )
        )
    return rows


def _classify(text: str) -> tuple[str | None, str | None, str | None, str | None]:
    section_match = SECTION_PATTERN.match(text)
    article_match = ARTICLE_PATTERN.match(text)
    paragraph_match = PARAGRAPH_PATTERN.match(text)
    section = section_match.group(0).replace(" ", "") if section_match else None
    article = article_match.group(1).replace(" ", "") if article_match else None
    title = article_match.group(2).strip() if article_match and article_match.group(2) else None
    paragraph = paragraph_match.group(1) if paragraph_match else None
    return section, article, title, paragraph


def _split_rows_by_length(rows: list[RowText], max_chars: int) -> list[list[RowText]]:
    groups: list[list[RowText]] = []
    current: list[RowText] = []
    for row in rows:
        if current and len(_join_rows(current + [row])) > max_chars:
            groups.append(current)
            current = []
        current.append(row)
    if current:
        groups.append(current)
    return groups


def _join_rows(rows: list[RowText]) -> str:
    return "\n".join(row.text for row in rows if row.text.strip()).strip()


def _make_chunk(
    document_id: str,
    sheet_id: str,
    sheet_name: str,
    rows: list[RowText],
    section: str | None,
    article: str | None,
    title: str | None,
    created_at: datetime,
    chunk_index: int,
) -> DocumentChunk:
    content = _join_rows(rows)
    min_row = min(row.min_row for row in rows)
    min_col = min(row.min_col for row in rows)
    max_row = max(row.max_row for row in rows)
    max_col = max(row.max_col for row in rows)
    cell_refs = tuple(ref for row in rows for ref in row.cell_refs)
    cell_start = f"{get_column_letter(min_col)}{min_row}"
    cell_end = f"{get_column_letter(max_col)}{max_row}"
    return DocumentChunk(
        id=f"{sheet_id}-C{chunk_index:04d}",
        document_id=document_id,
        sheet_id=sheet_id,
        sheet_name=sheet_name,
        cell_start=cell_start,
        cell_end=cell_end,
        cell_range=f"{cell_start}:{cell_end}",
        cell_refs=cell_refs,
        row_start=min_row,
        row_end=max_row,
        section=section,
        article=article,
        paragraph=None,
        title=title,
        content=content,
        chunk_index=chunk_index,
        content_hash=hashlib.sha256(content.encode("utf-8")).hexdigest(),
        created_at=created_at,
    )


def _with_chunk_index(chunk: DocumentChunk, chunk_index: int) -> DocumentChunk:
    return DocumentChunk(
        **{
            **chunk.__dict__,
            "id": f"{chunk.sheet_id}-C{chunk_index:04d}",
            "chunk_index": chunk_index,
        }
    )


def _cell_ref(cell: ParsedCell) -> str:
    return cell.merged_range if cell.merged_range else cell.coordinate


def _bounds_for_cells(cells: list[ParsedCell]) -> tuple[int, int, int, int]:
    min_row = min(cell.row_index for cell in cells)
    min_col = min(cell.column_index for cell in cells)
    max_row = max(cell.row_index for cell in cells)
    max_col = max(cell.column_index for cell in cells)
    for cell in cells:
        if not cell.merged_range:
            continue
        merged_min_col, merged_min_row, merged_max_col, merged_max_row = range_boundaries(cell.merged_range)
        min_row = min(min_row, merged_min_row)
        min_col = min(min_col, merged_min_col)
        max_row = max(max_row, merged_max_row)
        max_col = max(max_col, merged_max_col)
    return min_row, min_col, max_row, max_col
