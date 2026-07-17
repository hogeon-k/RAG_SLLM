from __future__ import annotations

from pathlib import Path
from zipfile import BadZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.services.exceptions import DocumentValidationError


class DocumentValidator:
    def __init__(self, max_file_size_bytes: int, allowed_extensions: tuple[str, ...] = (".xlsx",)) -> None:
        self._max_file_size_bytes = max_file_size_bytes
        self._allowed_extensions = tuple(ext.lower() for ext in allowed_extensions)

    def validate(self, source_path: Path) -> None:
        if not source_path.exists():
            raise DocumentValidationError("선택한 파일을 찾을 수 없습니다.")
        if not source_path.is_file():
            raise DocumentValidationError("일반 파일만 등록할 수 있습니다.")
        if source_path.suffix.lower() not in self._allowed_extensions:
            raise DocumentValidationError("현재 .xlsx 형식만 지원합니다.")

        file_size = source_path.stat().st_size
        if file_size == 0:
            raise DocumentValidationError("빈 파일은 등록할 수 없습니다.")
        if file_size > self._max_file_size_bytes:
            max_mb = self._max_file_size_bytes // (1024 * 1024)
            raise DocumentValidationError(f"파일 크기가 {max_mb}MB 제한을 초과했습니다.")
        if not is_zipfile(source_path):
            raise DocumentValidationError("엑셀 파일을 열 수 없습니다. 파일이 손상되었는지 확인해 주세요.")

        try:
            workbook = load_workbook(source_path, read_only=True, keep_links=False)
            try:
                if not workbook.sheetnames:
                    raise DocumentValidationError("시트가 없는 엑셀 파일은 등록할 수 없습니다.")
            finally:
                workbook.close()
        except DocumentValidationError:
            raise
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise DocumentValidationError("엑셀 파일을 열 수 없습니다. 파일이 손상되었는지 확인해 주세요.") from exc

