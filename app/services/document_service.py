from __future__ import annotations

import logging
import hashlib
import sqlite3
import uuid
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.config.settings import Settings
from app.models.document import Document, DocumentDeleteResult
from app.repositories.document_repository import DocumentRepository
from app.repositories.keyword_search_repository import KeywordSearchRepository
from app.repositories.search_index_repository import SearchIndexRepository
from app.services.document_validator import DocumentValidator
from app.services.exceptions import (
    DocumentDeleteError,
    DocumentRegistrationError,
    DocumentStorageError,
    DuplicateDocumentError,
)
from app.storage.file_storage import FileStorage
from app.storage.vector_storage import ChromaVectorRepository
from app.utils.hashing import calculate_sha256


class DocumentService:
    def __init__(
        self,
        settings: Settings,
        repository: DocumentRepository | None = None,
        storage: FileStorage | None = None,
        validator: DocumentValidator | None = None,
        logger: logging.Logger | None = None,
    ) -> None:
        self._settings = settings
        self._repository = repository or DocumentRepository(settings.database_path)
        self._storage = storage or FileStorage(settings.data_dir)
        self._validator = validator or DocumentValidator(
            settings.max_xlsx_bytes,
            settings.allowed_document_extensions,
        )
        self._logger = logger or logging.getLogger("rag_sllm")
        self._deleting_document_ids: set[str] = set()

    def status_message(self) -> str:
        return "엑셀과 규정 문서를 등록하고 관리하는 화면입니다."

    def register_document(
        self,
        source_path: Path,
        version: str | None = None,
        effective_date: date | None = None,
        revised_date: date | None = None,
        department: str | None = None,
    ) -> Document:
        source_path = Path(source_path)
        version = _clean_optional_text(version)
        department = _clean_optional_text(department)
        document_id = f"DOC-{uuid.uuid4()}"

        self._validator.validate(source_path)
        file_hash = calculate_sha256(source_path)
        duplicate = self._repository.get_by_hash(file_hash)
        if duplicate:
            raise DuplicateDocumentError(
                "동일한 내용의 문서가 이미 등록되어 있습니다.\n"
                f"등록 문서: {duplicate.original_name}"
            )
        content_duplicate = self._find_content_duplicate(source_path)
        if content_duplicate:
            raise DuplicateDocumentError(
                "동일한 내용의 문서가 이미 등록되어 있습니다.\n"
                f"등록 문서: {content_duplicate.original_name}"
            )

        stored_path = ""
        try:
            stored_path = self._storage.store_document(source_path, document_id, file_hash)
            document = Document(
                id=document_id,
                original_name=source_path.name,
                stored_path=stored_path,
                file_hash=file_hash,
                file_size_bytes=source_path.stat().st_size,
                version=version,
                effective_date=effective_date,
                revised_date=revised_date,
                department=department,
                is_latest=True,
                status="UPLOADED",
                error_message=None,
                uploaded_at=datetime.now().replace(microsecond=0),
            )
            return self._repository.create(document)
        except sqlite3.IntegrityError as exc:
            self._storage.cleanup_document(document_id)
            self._logger.warning("Duplicate document hash blocked for %s", source_path.name)
            raise DuplicateDocumentError("동일한 내용의 문서가 이미 등록되어 있습니다.") from exc
        except DocumentRegistrationError:
            if stored_path:
                self._storage.cleanup_document(document_id)
            raise
        except Exception as exc:
            self._storage.cleanup_document(document_id)
            self._logger.exception("Unexpected document registration failure for %s", source_path.name)
            raise DocumentStorageError("문서를 등록하는 중 예상하지 못한 오류가 발생했습니다.") from exc

    def list_documents(self) -> list[Document]:
        return self._repository.list_all()

    def delete_document(self, document_id: str) -> DocumentDeleteResult:
        document = self._repository.get_by_id(document_id)
        if document is None:
            raise DocumentDeleteError("DOCUMENT_NOT_FOUND", "선택한 문서를 찾을 수 없습니다.")
        if document.status in {"PARSING", "INDEXING"}:
            raise DocumentDeleteError("DOCUMENT_BUSY", "문서가 처리 중이어서 삭제할 수 없습니다.")
        if document_id in self._deleting_document_ids:
            raise DocumentDeleteError("DOCUMENT_DELETE_ALREADY_RUNNING", "이미 문서 삭제가 진행 중입니다.")

        self._deleting_document_ids.add(document_id)
        quarantine_dir = None
        vector_deleted_count = 0
        fts_deleted_count = 0
        internal_file_deleted = False
        try:
            self._storage.resolve(document.stored_path)
            quarantine_dir = self._storage.quarantine_document(document_id, document.stored_path)

            keyword_repository = KeywordSearchRepository(self._settings.database_path)
            fts_deleted_count = keyword_repository.count(document_id)

            index_repository = SearchIndexRepository(self._settings.database_path)
            index_status = index_repository.get(document_id)
            fingerprint = index_status.model_fingerprint if index_status and index_status.model_fingerprint else _delete_fingerprint(self._settings)
            vector_repository = ChromaVectorRepository(self._settings.vector_db_dir, self._settings.vector_collection, fingerprint)
            vector_deleted_count = vector_repository.count_document(document_id)
            vector_repository.delete_document(document_id)
            if vector_repository.count_document(document_id) != 0:
                raise DocumentDeleteError("VECTOR_DELETE_INCOMPLETE", "문서 검색 벡터를 완전히 삭제하지 못해 문서 삭제를 중단했습니다.")

            keyword_repository.remove_document(document_id)
            result = self._repository.delete_document_records(
                document_id,
                document.original_name,
                deleted_fts_count=fts_deleted_count,
                deleted_vector_count=vector_deleted_count,
                internal_file_deleted=False,
            )
            try:
                internal_file_deleted = self._storage.finalize_quarantine(quarantine_dir)
            except DocumentStorageError as exc:
                return DocumentDeleteResult(
                    document_id=result.document_id,
                    display_name=result.display_name,
                    deleted_document_count=result.deleted_document_count,
                    deleted_sheet_count=result.deleted_sheet_count,
                    deleted_cell_count=result.deleted_cell_count,
                    deleted_chunk_count=result.deleted_chunk_count,
                    deleted_fts_count=result.deleted_fts_count,
                    deleted_vector_count=result.deleted_vector_count,
                    internal_file_deleted=False,
                    history_preserved=True,
                    warning_code="FILE_FINALIZE_FAILED",
                )
            return DocumentDeleteResult(
                document_id=result.document_id,
                display_name=result.display_name,
                deleted_document_count=result.deleted_document_count,
                deleted_sheet_count=result.deleted_sheet_count,
                deleted_cell_count=result.deleted_cell_count,
                deleted_chunk_count=result.deleted_chunk_count,
                deleted_fts_count=result.deleted_fts_count,
                deleted_vector_count=result.deleted_vector_count,
                internal_file_deleted=internal_file_deleted,
                history_preserved=True,
                warning_code=None,
            )
        except DocumentDeleteError:
            self._restore_quarantine_after_failure(quarantine_dir, document_id)
            if vector_deleted_count or fts_deleted_count:
                self._mark_index_stale_after_partial_delete(document_id)
            raise
        except DocumentStorageError as exc:
            self._restore_quarantine_after_failure(quarantine_dir, document_id)
            raise DocumentDeleteError("INVALID_DOCUMENT_PATH", exc.user_message) from exc
        except Exception as exc:
            self._restore_quarantine_after_failure(quarantine_dir, document_id)
            if vector_deleted_count or fts_deleted_count:
                self._mark_index_stale_after_partial_delete(document_id)
            self._logger.exception("Unexpected document deletion failure for %s", document_id)
            raise DocumentDeleteError("INTERNAL_ERROR", "문서를 삭제하는 중 오류가 발생했습니다. 다른 문서와 질문 이력은 변경되지 않았습니다.") from exc
        finally:
            self._deleting_document_ids.discard(document_id)

    def set_lifecycle_status(self, document_id: str, lifecycle_status: str) -> Document:
        current = self._repository.get_by_id(document_id)
        if current is None:
            raise DocumentRegistrationError("문서를 찾을 수 없습니다.")
        if current.status in {"PARSING", "INDEXING"}:
            raise DocumentRegistrationError("처리 중인 문서는 업무 상태를 변경할 수 없습니다.")
        self._repository.update_lifecycle_status(document_id, lifecycle_status)
        updated = self._repository.get_by_id(document_id)
        if updated is None:
            raise DocumentRegistrationError("문서 상태 변경 후 문서를 다시 불러올 수 없습니다.")
        return updated

    def promote_current(self, document_id: str) -> tuple[Document, list[str]]:
        current = self._repository.get_by_id(document_id)
        if current is None:
            raise DocumentRegistrationError("문서를 찾을 수 없습니다.")
        if current.status in {"PARSING", "INDEXING"}:
            raise DocumentRegistrationError("처리 중인 문서는 현행 문서로 전환할 수 없습니다.")
        archived_ids = self._repository.promote_current(document_id)
        updated = self._repository.get_by_id(document_id)
        if updated is None:
            raise DocumentRegistrationError("문서 상태 변경 후 문서를 다시 불러올 수 없습니다.")
        return updated, archived_ids

    def update_version_metadata(
        self,
        document_id: str,
        *,
        version_label: str | None = None,
        effective_from: date | None = None,
        effective_to: date | None = None,
        document_family: str | None = None,
        supersedes_document_id: str | None = None,
    ) -> Document:
        current = self._repository.get_by_id(document_id)
        if current is None:
            raise DocumentRegistrationError("문서를 찾을 수 없습니다.")
        self._repository.update_version_metadata(
            document_id,
            version_label=version_label,
            effective_from=effective_from,
            effective_to=effective_to,
            document_family=document_family,
            supersedes_document_id=supersedes_document_id,
        )
        updated = self._repository.get_by_id(document_id)
        if updated is None:
            raise DocumentRegistrationError("문서 metadata 변경 후 문서를 다시 불러올 수 없습니다.")
        return updated

    def _restore_quarantine_after_failure(self, quarantine_dir, document_id: str) -> None:
        try:
            self._storage.restore_quarantine(quarantine_dir, document_id)
        except Exception:
            self._logger.exception("Failed to restore quarantined file for %s", document_id)

    def _mark_index_stale_after_partial_delete(self, document_id: str) -> None:
        try:
            SearchIndexRepository(self._settings.database_path).mark_stale(document_id)
        except Exception:
            self._logger.exception("Failed to mark partially deleted index stale for %s", document_id)

    def _find_content_duplicate(self, source_path: Path) -> Document | None:
        source_fingerprint = _calculate_workbook_content_fingerprint(source_path)
        for document in self._repository.list_all():
            stored_path = self._settings.data_dir / document.stored_path
            if not stored_path.exists():
                continue
            try:
                if _calculate_workbook_content_fingerprint(stored_path) == source_fingerprint:
                    return document
            except Exception:
                self._logger.warning("Stored workbook fingerprint check failed for %s", document.id, exc_info=True)
        return None


def _clean_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _delete_fingerprint(settings: Settings) -> str:
    return f"delete-only:{settings.embedding_model}"


def _calculate_workbook_content_fingerprint(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        digest = hashlib.sha256()
        for sheet in workbook.worksheets:
            digest.update(sheet.title.encode("utf-8"))
            digest.update(b"\0")
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    digest.update(f"{cell.coordinate}\t{cell.data_type}\t{cell.value}".encode("utf-8"))
                    digest.update(b"\0")
            digest.update(b"\1")
        return digest.hexdigest()
    finally:
        workbook.close()
